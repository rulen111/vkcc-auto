import os
from io import BytesIO

import requests
from openpyxl import load_workbook

from flask import (
    Blueprint, flash, redirect, render_template, request, send_file
)
from werkzeug.utils import secure_filename

# from vkclient import VKclient
# from wbhandler import WBHandler

ALLOWED_EXTENSIONS = {"xlsx"}


bp = Blueprint("filehandler", __name__)


class VKclient(object):
    """Defining a custom VK client to work with API requests.
    Contains an access token and a method for getting a specific photo album."""

    def __init__(self, access_token="", version="5.199"):
        self.access_token = access_token
        self.version = version
        self.auth_params = {
            "access_token": self.access_token,
            "v": self.version,
        }

    def get_short_link(self, url: str):
        """"""
        baseurl = "https://api.vk.com/method/utils.getShortLink"
        params = {
            "url": url,
            "private": 0,
        }

        # logging.info(f"User-{self.id}. Trying to get album '{album_id}'")
        response = requests.get(baseurl, params={**self.auth_params, **params})

        return response.json().get("response", {}).get("short_url", "")


class WBHandler(object):
    """"""

    def __init__(self, file):
        self.wb_file = file
        self.wb = load_workbook(filename=self.wb_file)
        self.active_ws = self.wb.active
        self.max_row = self.active_ws.max_row

    def get_link(self, row, col):
        value = self.active_ws.cell(row=row, column=col).value
        return value

    def write_new_link(self, row, col, new_link):
        cell = self.active_ws.cell(row=row, column=col)
        cell.value = new_link
        cell.hyperlink = new_link
        cell.style = "Hyperlink"

    def save(self, fp):
        self.wb.save(fp)


def allowed_file(filename):
    return "." in filename and \
        filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def payload(input_file, client_token, first_row=2, input_col=1, target_col=2):
    client = VKclient(client_token)
    wb = WBHandler(input_file)

    for row in range(first_row, wb.max_row + 1):
        link = wb.get_link(row, input_col)
        if not link:
            return wb
        short_link = client.get_short_link(link)
        wb.write_new_link(row, target_col, short_link)

    # virtual_workbook = BytesIO()
    # wb.save(virtual_workbook)
    # virtual_workbook.seek(0)
    #
    # return virtual_workbook
    return wb


@bp.route("/", methods=("GET", "POST"))
def index():
    if request.method == "POST":
        if "file" not in request.files:
            flash("No file part")

            return redirect(request.url)

        file = request.files["file"]
        if file.filename == "":
            flash("No selected file")

            return redirect(request.url)

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            # virtual_workbook = payload(file, TOKEN)

            # wb = WBHandler(file)
            wb = payload(file, TOKEN)
            virtual_workbook = BytesIO()
            wb.save(virtual_workbook)
            virtual_workbook.seek(0)

            return send_file(virtual_workbook, as_attachment=True, download_name=f"{filename}_output.xlsx",
                             mimetype="application/vnd.ms-excel")

    return render_template("index.html")
