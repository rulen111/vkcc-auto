from openpyxl import load_workbook


class WBHandler(object):
    """"""

    def __init__(self, file: str):
        self.wb_file = file
        self.wb = load_workbook(filename=self.wb_file)
        self.active_ws = self.wb.active
        self.max_row = self.active_ws.max_row

    def get_link(self, row, col):
        value = self.active_ws.cell(row=row, column=col).value
        return value

    def write_new_link(self, row, col, new_link):
        cell = self.active_ws.cell(row=row, column=col)
        cell.value = new_link
        cell.hyperlink = new_link
        cell.style = "Hyperlink"

    def save(self, fp):
        self.wb.save(fp)
