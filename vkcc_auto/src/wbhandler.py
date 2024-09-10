from typing import IO
import os

from openpyxl import load_workbook
from openpyxl.styles import Font
from werkzeug.datastructures import FileStorage

from .utils import WBReaderError, WBWriterError


class WBHandler(object):
    """
    Openpyxl workbook wrapper, handles workbook operations
    """
    def __init__(self, file: str | os.PathLike | IO | FileStorage) -> None:
        """
        Load excel workbook to start working
        :param file: excel workbook file or path
        """
        self.wb_file = file
        try:
            self.wb = load_workbook(filename=self.wb_file)
            self.active_ws = self.wb.active
            self.max_row = self.active_ws.max_row
        except Exception:
            raise WBReaderError()

    def get_value(self, row: int, col: int) -> str:
        """
        Get str value of specified cell
        :param row: cell row number (starts from 1)
        :param col: cell col number (starts from 1)
        :return: value of specified cell
        """
        return self.active_ws.cell(row=row, column=col).value

    def write_header(self, row: int, col: int, value: str = "Short link") -> None:
        """
        Write header for target column with specified value
        :param row: cell row number (starts from 1)
        :param col: cell col number (starts from 1)
        :param value: text to write in cell
        """
        cell = self.active_ws.cell(row=row, column=col)
        cell.value = value
        cell.font = Font(bold=True)

    def write_value(self, row: int, col: int, value: str) -> None:
        """
        Writes new value (in HyperLink style) to specified cell
        :param row: cell row number (starts from 1)
        :param col: cell col number (starts from 1)
        :param value: url to write
        """
        cell = self.active_ws.cell(row=row, column=col)
        cell.value = value
        cell.hyperlink = value
        cell.style = "Hyperlink"

    def save(self, file: str | os.PathLike | IO) -> None:
        """
        Save workbook to file or stream
        :param file: excel workbook file or path
        """
        try:
            self.wb.save(file)
        except Exception:
            raise WBWriterError()
